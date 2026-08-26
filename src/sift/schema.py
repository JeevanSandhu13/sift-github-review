"""Sift — schema extractor.

Produces a structural summary of a dataset (variable names, types, labels,
observation count, optionally NA counts / distinct counts) from files on
the researcher's machine. Supported formats: `.dta` (Stata), `.rds` (R),
`.csv`, `.tsv`, `.parquet`, `.jsonl` / `.ndjson`.

No individual observation values are ever returned. The only potentially
disclosive pieces of output are:
- Variable labels (short human-readable strings; may contain sensitive
  descriptions like "Primary diagnosis" — exposed at depth >=
  names_types_labels).
- Value labels, i.e. the level-name dictionary for a categorical variable
  (e.g. `{1: "Control", 2: "Treatment"}` — exposed at depth >=
  names_types_labels, same tier as variable labels).

Neither is a per-observation value; both are metadata attached to the
variable. The researcher can dial depth down if their column labels or
value labels are themselves sensitive. Step 5 (real sanitizer with SDC
rules) will add per-variable controls.

Depths (graded from conservative → permissive):
  names_only
  names_types
  names_types_labels
  names_types_labels_summary     (+ NA count, distinct count for categoricals)

The ``na_count`` field at the summary depth is subject to primary
cell suppression (see ``_suppress_rare_count``): a count below the
threshold on the rarer side is replaced with a ``<N`` marker, same
shape as :func:`sift.sdc.suppression_marker`. Without this, a
column with exactly one missing value would re-identify that
observation through ``get_schema`` before the stricter
``request_data`` / result-sanitizer paths ever ran.

Not included, ever, at any depth:
- Actual observation values
- Min / max / mean / quantiles on numerics (these are individual values)
- Frequency tables (step 5 territory — requires SDC cell-suppression)
"""

from __future__ import annotations

import os
import zipfile
import zlib
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Literal

from sift.limits import ZIP_CONTAINER_MAX_MEMBERS
from sift.text_safety import safe_keys_sequence, safe_text


Depth = Literal[
    "names_only",
    "names_types",
    "names_types_labels",
    "names_types_labels_summary",
]

_VALID_DEPTHS: frozenset[str] = frozenset(
    ("names_only", "names_types", "names_types_labels", "names_types_labels_summary")
)


# Centralised data-file extension allowlist. Imported by every module
# that scans a session dir for "the researcher's datasets" (the bridge,
# session_state, drop-zone hint copy, file dialog filters). Adding a
# new format means: (a) add the extension here, (b) add a dispatch
# branch in ``extract()``/``load_data()`` below, (c) make sure any
# parsing dependency is in pyproject.toml.
#
# ``.jsonl`` and ``.ndjson`` are aliases — pandas reads both with
# ``read_json(lines=True)`` and researchers ship under either name.
DATA_EXTENSIONS: tuple[str, ...] = (
    ".csv",
    ".dta",
    ".rds",
    # R workspaces. Unlike .rds these may contain several objects; Sift
    # accepts them only when exactly one data frame is present so dataset
    # selection can never happen silently.
    ".rda",
    ".rdata",
    ".parquet",
    # Apache Arrow IPC / Feather and ORC: common zero-copy exchange
    # formats in genomics, geospatial, Spark, and data engineering.
    ".feather",
    ".arrow",
    ".ipc",
    ".orc",
    ".jsonl",
    ".ndjson",
    ".json",
    ".tsv",
    # SPSS. The lingua franca of psychology, sociology, survey
    # research and much of public health. ``.zsav`` is the same
    # format zlib-compressed; pyreadstat reads both via read_sav.
    ".sav",
    ".zsav",
    ".por",
    # SAS. ``.sas7bdat`` is the native dataset format of pharma,
    # clinical trials and several national statistics offices;
    # ``.xpt`` (SAS Transport) is the FDA submission format.
    ".sas7bdat",
    ".xpt",
    # Excel. The most common data container in existence. A saved
    # worksheet selection is honored; otherwise the first worksheet
    # is read. The header row is assumed by the extractor.
    ".xlsx",
    # Legacy binary Excel remains common in government and long-running
    # longitudinal studies; ODS is the open-standard spreadsheet used by
    # LibreOffice and many Linux research desktops.
    ".xls",
    ".ods",
    # Gzip is accepted only when the complete name ends in one of
    # .csv.gz, .tsv.gz, .jsonl.gz, or .ndjson.gz. Dispatch fails closed for
    # every other gzip payload and checks decompressed expansion first.
    ".gz",
    # Explicit-selection/container workflows. These extensions are recognized
    # by the desktop but are materialized to a safe local table through
    # sift.format_selection before ordinary schema/analysis loading.
    ".zip", ".avro", ".xml", ".dbf",
    ".h5", ".hdf5", ".nc", ".netcdf", ".mat",
    ".fits", ".fit", ".fts",
    ".geojson", ".gpkg", ".shp", ".tif", ".tiff", ".vrt",
    ".vcf", ".bcf", ".bed",
    ".nii", ".dcm", ".fhir",
)

# Our coarse taxonomy. Step 4+ analysis-result schemas consume these; keep
# the vocabulary small and stable.
_TYPE_NUMERIC = "numeric"
_TYPE_INTEGER = "integer"
_TYPE_CATEGORICAL = "categorical"
_TYPE_STRING = "string"
_TYPE_BOOLEAN = "boolean"
_TYPE_DATETIME = "datetime"
_TYPE_UNKNOWN = "unknown"


# Primary cell-suppression threshold for schema summary metadata. The
# value here mirrors :class:`sift.sanitizer.SDCConfig.cell_suppression_threshold`
# (10) but is kept inline so this module doesn't take a runtime
# dependency on the regression sanitizer. Schema summary publishes
# ``na_count`` per variable at the richest depth tier; without
# suppression a column with exactly one missing value (or one present
# value, in a mostly-empty column) would re-identify that observation
# directly — the same disclosive concern that primary cell
# suppression solves for frequency tables. ``request_data`` /
# regression-result paths apply their own SDC, but ``get_schema`` is
# the first surface the model can call against a dataset and runs
# before either of those, so the suppression has to live here too.
_SCHEMA_SUMMARY_THRESHOLD = 10


@contextmanager
def _open_arrow_ipc(path: Path):
    """Open Arrow IPC in either file (footer) or streaming form."""
    import pyarrow as pa
    import pyarrow.ipc as ipc

    source = pa.memory_map(str(path), "r")
    try:
        try:
            reader = ipc.open_file(source)
            streaming = False
        except (pa.ArrowInvalid, OSError):
            source.seek(0)
            reader = ipc.open_stream(source)
            streaming = True
        yield reader, streaming
    finally:
        source.close()


def _arrow_batches(reader: Any, streaming: bool):
    if streaming:
        yield from reader
    else:
        for index in range(reader.num_record_batches):
            yield reader.get_batch(index)


def _suppress_rare_count(value: int, n: int, threshold: int) -> int | str:
    """Return ``value`` unchanged when it sits comfortably above the
    threshold on both sides; otherwise return the suppression marker.

    "Both sides" is the symmetric edge case: a column with
    ``na_count == 1`` identifies the one missing observation; a
    column with ``n - na_count == 1`` identifies the one present
    observation. Either is a re-identification channel, so suppress
    when the rarer side falls below ``threshold``. ``value == 0`` and
    ``value == n`` are both safe (no rare subgroup) and pass
    through.
    """
    if value < 0 or n < 0 or threshold <= 0:
        return value
    rarer = min(value, n - value) if n >= value else value
    if rarer == 0:
        return value
    if rarer < threshold:
        # ``<10``-style marker, same shape as
        # :func:`sift.sdc.suppression_marker`.
        return f"<{threshold}"
    return value



# ---------------------------------------------------------------------------
# Full-load size guard
# ---------------------------------------------------------------------------
#
# ``load_data`` materialises an ENTIRE dataset as a pandas DataFrame.
# In-memory pandas representation routinely runs 3-10x the on-disk size
# (object-dtype string columns are the worst case: every cell becomes a
# separate Python object). Uploads are capped at 1 GB per file, so an
# unguarded full load can ask for ~10 GB of RAM on a laptop with 8 —
# the OS starts swapping, the UI beachballs, and in the worst case the
# app is OOM-killed mid-session.
#
# The paths that reach a full load are narrow and all optional:
#   - ``request_data`` (bounded facts about ONE variable)
#   - the ``.rds`` row-count fallback (no metadata-only reader exists)
#   - summary-depth schema extraction
# None of them is worth an OOM. Above the ceiling we raise a typed
# error whose message tells the researcher AND the model what to do
# instead: run the question as a script, where the analysis runtime
# streams/chunks on its own terms inside the sandbox and only a
# sanitized aggregate crosses back.
#
# The ceiling is on-disk bytes (cheap to check, no read required) and
# is deliberately generous: 512 MB on disk is a large survey/admin
# extract and still loads comfortably on a 16 GB machine. Researchers
# with big machines can raise it; the env var is read at call time so
# it can be changed without a restart.
_DEFAULT_FULL_LOAD_MAX_BYTES = 512 * 1024 * 1024


def full_load_max_bytes() -> int:
    """Return the current full-load ceiling in bytes.

    Read at call time (not import time) so a researcher can raise the
    limit via ``SIFT_MAX_LOAD_BYTES`` without restarting. Invalid or
    non-positive values fall back to the default rather than
    disabling the guard — a typo must never silently remove a memory
    protection.
    """
    raw = os.environ.get("SIFT_MAX_LOAD_BYTES", "").strip()
    if not raw:
        return _DEFAULT_FULL_LOAD_MAX_BYTES
    try:
        val = int(raw)
    except ValueError:
        return _DEFAULT_FULL_LOAD_MAX_BYTES
    return val if val > 0 else _DEFAULT_FULL_LOAD_MAX_BYTES


# Zip-container formats Sift reads whole. A malicious (or just
# badly-behaved) member table can report a huge UNCOMPRESSED size
# while itself being tiny on disk — the classic "zip bomb" shape.
# Any format added here must expose an uncompressed-size figure via
# ``zipfile`` without decompressing member data (true of every
# format below: .xlsx is a plain zip of XML parts).
_ZIP_CONTAINER_EXTENSIONS = frozenset({".xlsx", ".ods"})
_EXPLICIT_SELECTION_EXTENSIONS = frozenset({
    ".zip", ".avro", ".xml", ".dbf", ".h5", ".hdf5", ".nc", ".netcdf",
    ".mat", ".fits", ".fit", ".fts", ".geojson", ".gpkg", ".shp",
    ".tif", ".tiff", ".vrt", ".vcf", ".vcf.gz", ".bcf", ".bed",
    ".nii", ".nii.gz", ".dcm", ".fhir",
})


def _logical_suffix(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix != ".gz":
        return suffix
    inner = Path(path.stem).suffix.lower()
    if inner not in {".csv", ".tsv", ".jsonl", ".ndjson", ".nii", ".vcf"}:
        raise SchemaExtractError(
            "gzip input must be named .csv.gz, .tsv.gz, .jsonl.gz, or .ndjson.gz"
        )
    return f"{inner}.gz"


def _guard_gzip_bomb(path: Path, ceiling: int) -> None:
    import gzip

    total = 0
    with gzip.open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(256 * 1024), b""):
            total += len(chunk)
            if total > ceiling:
                raise DatasetTooLargeError(
                    "gzip dataset expands beyond Sift's safe full-load ceiling"
                )

# A zip with an absurd number of entries is its own resource-exhaustion
# vector independent of any single member's size (iterating the
# central directory, opening/closing thousands of tiny members).
# Legitimate .xlsx files have on the order of tens to low hundreds of
# parts (sheets, styles, shared strings, media, relationships); this
# cap is generous headroom above that while still bounding pathological
# cases.
_ZIP_MAX_MEMBERS = ZIP_CONTAINER_MAX_MEMBERS


def _guard_zip_bomb(dataset_path: Path, ceiling: int) -> None:
    """Refuse a zip-container file whose DECLARED uncompressed size
    (or member count) would blow the memory ceiling once unpacked,
    even though the file is small on disk.

    Verified empirically during development: a ~104 MB .xlsx built
    from a highly-compressible worksheet decompresses to ~1.4 GB of
    XML (a 14x expansion from a deliberately modest PoC — real zip
    bombs using nested archives or more redundant payloads routinely
    exceed 1000x) while sailing under the 512 MB on-disk ceiling
    that ``_guard_full_load`` checks. ``zipfile``'s central-directory
    read used here is cheap and safe: ``ZipInfo.file_size`` is a
    stored field read from the archive's directory listing, not a
    result of decompressing anything, so this check itself cannot be
    used to mount the same class of attack it defends against.

    Malformed/non-zip files (a mislabelled extension, truncated
    upload) are NOT this function's concern — they raise
    ``BadZipFile`` here, which the caller lets propagate up to the
    normal "file failed to parse" error path, same as any other
    corrupt-file failure mode.
    """
    with zipfile.ZipFile(dataset_path) as zf:
        infos = zf.infolist()
        if len(infos) > _ZIP_MAX_MEMBERS:
            raise DatasetTooLargeError(
                f"{dataset_path.name} contains {len(infos)} internal "
                f"parts, above the {_ZIP_MAX_MEMBERS}-part safety "
                f"ceiling for a spreadsheet file — this doesn't look "
                f"like a normal .xlsx. Refusing to open it."
            )
        total_uncompressed = sum(i.file_size for i in infos)
        if total_uncompressed > ceiling:
            on_disk = dataset_path.stat().st_size
            raise DatasetTooLargeError(
                f"{dataset_path.name} is only "
                f"{on_disk // (1024 * 1024)} MB on disk, but its "
                f"internal contents decompress to "
                f"{total_uncompressed // (1024 * 1024)} MB — above "
                f"the {ceiling // (1024 * 1024)} MB ceiling for "
                f"loading a dataset fully into memory. This is either "
                f"a genuinely enormous spreadsheet or a malformed / "
                f"maliciously crafted file; Sift refuses to unpack it. "
                f"Raise SIFT_MAX_LOAD_BYTES if this machine has the RAM "
                f"and you trust the source."
            )


def _guard_rds_bomb(dataset_path: Path, ceiling: int) -> None:
    """Refuse a gzip-compressed .rds file whose DECOMPRESSED size would
    blow the memory ceiling, even though the file is small on disk.

    R's ``saveRDS()`` defaults to gzip compression (``compress = TRUE``
    maps to gzip), so this covers the common case a researcher will
    actually hit. Verified empirically during development: a 300k-row
    data frame of two highly-repetitive string columns produced a
    1.4 MB gzip-compressed .rds that decompresses to ~29 MB (a ~21x
    expansion) — the same "small on disk, huge unpacked" shape
    ``_guard_zip_bomb`` defends against for .xlsx, and RDS's flat
    binary serialization (no per-row XML tag overhead diluting
    redundancy the way .xlsx's markup does) can plausibly compress
    far better than that for a deliberately crafted file.

    Deliberately does NOT trust gzip's trailer ISIZE field (the
    declared uncompressed size) the way ``_guard_zip_bomb`` trusts
    ``zipfile``'s central directory: ISIZE is stored mod 2**32, so a
    file whose true decompressed size exceeds 4 GB can report an
    arbitrary smaller value there — trivial for an attacker to
    engineer by padding to a chosen residue, and 4 GB is well within
    reach of a small, highly-redundant .rds. Instead this does REAL
    bounded work: stream-decompress in fixed chunks through
    ``zlib.decompressobj``, using its ``max_length`` argument to cap
    output per call, and abort the instant total output crosses
    ``ceiling`` — so the check can never be fooled by a crafted
    trailer, and never does more than ~``ceiling`` bytes of actual
    decompression even against a file that would truly unpack to
    gigabytes.

    Non-gzip .rds files (uncompressed, or the rarer ``compress =
    "bzip2"`` / ``"xz"`` options) are not this function's concern —
    gzip is R's default and by far the common case, and the on-disk-
    size check in ``_guard_full_load`` still applies to every .rds
    regardless of compression scheme. A malformed/truncated gzip
    stream is also not this function's concern: it raises
    ``zlib.error``, which is treated as "nothing more to check here"
    and left for the real reader (``pyreadr``) to fail on normally.
    """
    try:
        with open(dataset_path, "rb") as f:
            magic = f.read(2)
    except OSError:
        return
    if magic != b"\x1f\x8b":
        return  # not gzip — on-disk-size check is all we can cheaply do
    decompressor = zlib.decompressobj(wbits=zlib.MAX_WBITS | 16)
    total_out = 0
    read_chunk = 256 * 1024
    try:
        with open(dataset_path, "rb") as f:
            pending = b""
            while True:
                if not pending:
                    pending = f.read(read_chunk)
                    if not pending:
                        break
                out = decompressor.decompress(pending, ceiling + 1 - total_out)
                total_out += len(out)
                pending = decompressor.unconsumed_tail
                if total_out > ceiling:
                    on_disk = dataset_path.stat().st_size
                    raise DatasetTooLargeError(
                        f"{dataset_path.name} is only "
                        f"{on_disk // (1024 * 1024)} MB on disk, but its "
                        f"gzip-compressed contents decompress to over "
                        f"{ceiling // (1024 * 1024)} MB — above the "
                        f"ceiling for loading a dataset fully into "
                        f"memory. This is either a genuinely enormous R "
                        f"object or a malformed / maliciously crafted "
                        f"file; Sift refuses to unpack it. Raise "
                        f"SIFT_MAX_LOAD_BYTES if this machine has the "
                        f"RAM and you trust the source."
                    )
    except zlib.error:
        return  # malformed gzip stream — let pyreadr's own error surface


def _guard_zsav_bomb(dataset_path: Path, ceiling: int) -> None:
    """Refuse a compressed SPSS .zsav file whose declared row/column
    dimensions imply a decompressed size beyond the memory ceiling.

    .zsav (SPSS's zlib-compressed SAV variant) stores its data in
    zlib-compressed blocks referenced from a trailer index — unlike
    gzip there is no single well-known trailer field for total
    uncompressed size, and hand-parsing IBM/ReadStat's block-index
    format to extract one would be exactly the kind of fragile,
    easy-to-get-subtly-wrong logic that is worse than no check at
    all. Instead this uses pyreadstat's ``metadataonly=True`` read —
    already the codebase's established fast path for row counts (see
    ``_audit_row_count``) and confirmed not to touch the compressed
    data blocks — to read the declared row count and per-column
    storage widths from the file's (uncompressed) dictionary header,
    and multiplies them out. SPSS stores every value at its declared
    fixed width regardless of actual content, so ``rows * sum(column
    widths)`` IS the file's real uncompressed data size, not just an
    estimate: verified empirically against a matching plain
    (uncompressed) .sav built from the same 300k-row data — computed
    24,000,000 bytes vs. 24,000,703 bytes actually on disk, matching
    to within header overhead.
    """
    try:
        import pyreadstat
        _df, meta = pyreadstat.read_sav(str(dataset_path), metadataonly=True)
    except Exception:
        return  # let the real reader's error surface normally
    n_rows = meta.number_rows
    widths = meta.variable_storage_width
    if n_rows is None or not widths:
        return
    estimated_bytes = int(n_rows) * sum(widths.values())
    if estimated_bytes > ceiling:
        on_disk = dataset_path.stat().st_size
        raise DatasetTooLargeError(
            f"{dataset_path.name} is only {on_disk // (1024 * 1024)} MB "
            f"on disk, but its declared {n_rows:,} rows x {len(widths)} "
            f"columns imply roughly "
            f"{estimated_bytes // (1024 * 1024)} MB decompressed — "
            f"above the {ceiling // (1024 * 1024)} MB ceiling for "
            f"loading a dataset fully into memory. This is either a "
            f"genuinely enormous SPSS file or a malformed / "
            f"maliciously crafted one; Sift refuses to unpack it. "
            f"Raise SIFT_MAX_LOAD_BYTES if this machine has the RAM "
            f"and you trust the source."
        )


def _guard_columnar_expansion(
    dataset_path: Path,
    ceiling: int,
    columns: list[str] | None = None,
) -> None:
    """Bound compressed Arrow-family files using footer dimensions.

    Parquet exposes an exact uncompressed byte total. Other Arrow-family
    containers expose rows and columns, for which 64 bytes per cell is a
    deliberately conservative materialisation estimate (pandas object
    columns commonly exceed it). Any metadata failure is left to the real
    parser; this helper never turns corruption into a false success.
    """
    suffix = dataset_path.suffix.lower()
    estimated = 0
    try:
        if suffix == ".parquet":
            import pyarrow.parquet as pq

            metadata = pq.ParquetFile(str(dataset_path)).metadata
            if columns is None:
                estimated = sum(
                    metadata.row_group(index).total_byte_size
                    for index in range(metadata.num_row_groups)
                )
            else:
                wanted = set(columns)
                estimated = sum(
                    group.column(column_index).total_uncompressed_size
                    for group_index in range(metadata.num_row_groups)
                    for group in (metadata.row_group(group_index),)
                    for column_index in range(group.num_columns)
                    if group.column(column_index).path_in_schema in wanted
                )
        elif suffix == ".feather":
            import pyarrow.feather as feather

            rows = feather.read_table(str(dataset_path), columns=[]).num_rows
            import pyarrow as pa
            import pyarrow.ipc as ipc

            with pa.memory_map(str(dataset_path), "r") as source:
                schema_columns = len(ipc.open_file(source).schema)
            width = len(columns) if columns is not None else schema_columns
            estimated = rows * max(1, width) * 64
        elif suffix in (".arrow", ".ipc"):
            with _open_arrow_ipc(dataset_path) as (reader, streaming):
                rows = sum(
                    batch.num_rows for batch in _arrow_batches(reader, streaming)
                )
                width = len(columns) if columns is not None else len(reader.schema)
                estimated = rows * max(1, width) * 64
        elif suffix == ".orc":
            import pyarrow.orc as orc

            reader = orc.ORCFile(str(dataset_path))
            width = len(columns) if columns is not None else len(reader.schema)
            estimated = reader.nrows * max(1, width) * 64
    except Exception:  # noqa: BLE001 - normal parser surfaces details later
        return
    if estimated > ceiling:
        raise DatasetTooLargeError(
            f"{dataset_path.name}'s columnar metadata implies roughly "
            f"{estimated // (1024 * 1024)} MB when materialized, above "
            f"the {ceiling // (1024 * 1024)} MB full-load ceiling. Read "
            f"it in a sandboxed script with Arrow/DuckDB projection or "
            f"raise SIFT_MAX_LOAD_BYTES if this machine has the RAM."
        )


def _guard_full_load(
    dataset_path: Path, columns: list[str] | None = None,
) -> None:
    """Refuse a full in-memory load of an oversized dataset.

    Checks on-disk size for every format, and — for formats that wrap
    a compressed container (.xlsx via ``_guard_zip_bomb``, gzip-
    compressed .rds via ``_guard_rds_bomb``, .zsav via
    ``_guard_zsav_bomb``) — ALSO checks a declared/estimated
    uncompressed size, since on-disk size alone is meaningless for a
    compressed container. Every check here runs before any library
    touches the file's actual content.
    """
    try:
        size = dataset_path.stat().st_size
    except OSError:
        return  # can't stat → let the normal load path report the error
    ceiling = full_load_max_bytes()
    suffix = _logical_suffix(dataset_path)
    projected_columnar = columns is not None and suffix in {
        ".parquet", ".feather", ".arrow", ".ipc", ".orc",
    }
    if size > ceiling and not projected_columnar:
        raise DatasetTooLargeError(
            f"{dataset_path.name} is {size // (1024 * 1024)} MB on disk, "
            f"above the {ceiling // (1024 * 1024)} MB ceiling for loading "
            f"a dataset fully into memory (in-memory size is typically "
            f"several times the file size). Run this as a script instead "
            f"— the analysis runtime can read the file in chunks inside "
            f"the sandbox and only a disclosure-controlled summary comes "
            f"back. Raise SIFT_MAX_LOAD_BYTES if this machine has the RAM."
        )
    if suffix.endswith(".gz"):
        _guard_gzip_bomb(dataset_path, ceiling)
    elif suffix in _ZIP_CONTAINER_EXTENSIONS:
        _guard_zip_bomb(dataset_path, ceiling)
    elif suffix in {".rds", ".rda", ".rdata"}:
        _guard_rds_bomb(dataset_path, ceiling)
    elif suffix == ".zsav":
        _guard_zsav_bomb(dataset_path, ceiling)
    elif suffix in (".parquet", ".feather", ".arrow", ".ipc", ".orc"):
        _guard_columnar_expansion(dataset_path, ceiling, columns=columns)


def load_data(
    dataset_path: Path,
    *,
    sheet: str | int | None = None,
    columns: list[str] | None = None,
    filters: list[tuple[str, str, Any]] | None = None,
    r_object: str | None = None,
) -> Any:
    """Load the dataset at `dataset_path` as a pandas DataFrame.

    Dispatches by extension (.dta / .rds / .csv). Used by
    ``extract()`` at the summary depth, and by ``data_request`` for
    computing bounded sanitized facts about a variable.

    ``sheet`` selects a worksheet for ``.xlsx`` (name or 0-based
    index); ignored for every other format. ``None`` preserves the
    long-standing default of the first worksheet — existing callers
    that never pass this argument see byte-identical behaviour.

    ``columns`` is a column-projection hint: when the format supports
    reading a subset of columns cheaply (currently only ``.parquet``,
    via pyarrow), only those columns are read from disk. ``None``
    (the default) reads every column, exactly as before this
    parameter existed. Ignored for formats with no partial-column
    reader — the caller always gets back every column in that case,
    never a silent partial result.

    Returns the full data as a DataFrame. Callers are responsible for
    never returning raw values — in Sift, only ``data_request``
    and schema extraction load data, and both sanitize before emitting.
    """
    import pandas as pd

    suffix = _logical_suffix(dataset_path)
    if filters is not None and suffix != ".parquet":
        raise SchemaExtractError(
            "predicate pushdown is supported only for Parquet datasets; "
            "refusing to ignore filters for this format"
        )
    if suffix in _EXPLICIT_SELECTION_EXTENSIONS:
        raise SchemaExtractError(
            f"{suffix} requires an explicit object/variable/layer selection; "
            "materialize it through Sift's isolated format-selection workflow first"
        )
    _guard_full_load(dataset_path, columns=columns)
    if suffix == ".dta":
        import pyreadstat
        df, _meta = pyreadstat.read_dta(str(dataset_path))
        return df
    if suffix in (".sav", ".zsav", ".por"):
        import pyreadstat
        reader = pyreadstat.read_por if suffix == ".por" else pyreadstat.read_sav
        df, _meta = reader(str(dataset_path))
        return df
    if suffix == ".sas7bdat":
        import pyreadstat
        df, _meta = pyreadstat.read_sas7bdat(str(dataset_path))
        return df
    if suffix == ".xpt":
        import pyreadstat
        df, _meta = pyreadstat.read_xport(str(dataset_path))
        return df
    if suffix in (".xlsx", ".xls", ".ods"):
        # First worksheet by default, header row assumed — same scope
        # as the schema extractor so the two views can't disagree.
        # ``sheet`` lets a caller (currently: the researcher's saved
        # per-dataset selection in policy.py) read a different one.
        return pd.read_excel(
            dataset_path, sheet_name=(0 if sheet is None else sheet),
        )
    if suffix in {".rds", ".rda", ".rdata"}:
        return _read_single_r_dataframe(dataset_path, object_name=r_object)
    if suffix == ".csv":
        # Use the same header peek as ``_extract_csv`` / ``row_count``
        # so a headerless numeric CSV (``1,2,3\n4,5,6``) doesn't get
        # row 1 consumed as the column header here while the schema
        # surface reports positional names (``0,1,2…``) plus 2 rows.
        # The disagreement broke ``request_data``: the model picked
        # variable ``"0"`` from the schema response, then ``load_data``
        # gave it columns ``["1","2","3"]`` and 1 data row.
        enc, sep, dec = text_table_params(dataset_path, ".csv")
        has_header = _csv_has_header(dataset_path, sep)
        return pd.read_csv(
            dataset_path, sep=sep, encoding=enc, decimal=dec,
            header=0 if has_header else None, low_memory=False,
        )
    if suffix == ".tsv":
        enc, sep, dec = text_table_params(dataset_path, ".tsv")
        has_header = _csv_has_header(dataset_path, sep)
        return pd.read_csv(
            dataset_path, sep=sep, encoding=enc, decimal=dec,
            header=0 if has_header else None, low_memory=False,
        )
    if suffix in {".csv.gz", ".tsv.gz"}:
        delimiter = "\t" if suffix == ".tsv.gz" else ","
        has_header = _gzip_has_header(dataset_path, delimiter)
        return pd.read_csv(
            dataset_path, sep=delimiter, compression="gzip",
            header=0 if has_header else None, low_memory=False,
        )
    if suffix == ".parquet":
        # pandas dispatches to pyarrow (preferred) or fastparquet —
        # pyarrow is a declared dep of sift so this works out of
        # the box. Parquet preserves dtypes so column types come
        # back exactly as the writer set them. ``columns=None`` is
        # pandas's own "read every column" default, so passing it
        # through unconditionally is a no-op for every caller that
        # doesn't ask for projection.
        return pd.read_parquet(dataset_path, columns=columns, filters=filters)
    if suffix == ".feather":
        return pd.read_feather(dataset_path, columns=columns)
    if suffix in (".arrow", ".ipc"):
        with _open_arrow_ipc(dataset_path) as (reader, _streaming):
            table = reader.read_all()
        if columns is not None:
            table = table.select(columns)
        return table.to_pandas()
    if suffix == ".orc":
        import pyarrow.orc as orc

        return orc.ORCFile(str(dataset_path)).read(columns=columns).to_pandas()
    if suffix in (".jsonl", ".ndjson"):
        # Line-delimited JSON: one record per line. Top-level JSON
        # arrays of arbitrary shape are NOT supported (each row
        # would have to be a flat object for the schema extractor
        # to make sense of it); researchers can convert with `jq`
        # if needed.
        return pd.read_json(dataset_path, lines=True)
    if suffix in {".jsonl.gz", ".ndjson.gz"}:
        return pd.read_json(dataset_path, lines=True, compression="gzip")
    if suffix == ".json":
        # Standard JSON record arrays / pandas-compatible table objects.
        # Unlike JSONL, this is necessarily a bounded full-file parse.
        return pd.read_json(dataset_path)
    raise SchemaExtractError(
        f"unsupported format: {suffix!r}. Sift reads "
        ".dta, .rds, .csv, .tsv, .parquet, .feather, .arrow/.ipc, "
        ".orc, .json, .jsonl, .ndjson."
    )


def _arrow_filter_expression(predicates: list[tuple[str, str, Any]] | None):
    """Compile a bounded conjunction without evaluating observations."""
    if not predicates:
        return None
    if len(predicates) > 64:
        raise SchemaExtractError("no more than 64 pushed-down predicates are allowed")
    import pyarrow.dataset as ds

    expression = None
    for predicate in predicates:
        if not isinstance(predicate, tuple) or len(predicate) != 3:
            raise SchemaExtractError("each predicate must be (column, operator, value)")
        column, operator, value = predicate
        if not isinstance(column, str) or not column or len(column) > 512:
            raise SchemaExtractError("predicate column is invalid")
        field = ds.field(column)
        if operator == "==":
            current = field == value
        elif operator == "!=":
            current = field != value
        elif operator == "<":
            current = field < value
        elif operator == "<=":
            current = field <= value
        elif operator == ">":
            current = field > value
        elif operator == ">=":
            current = field >= value
        elif operator in {"in", "not in"}:
            if not isinstance(value, (list, tuple)) or len(value) > 10_000:
                raise SchemaExtractError("membership predicate values must be bounded")
            current = field.isin(list(value))
            if operator == "not in":
                current = ~current
        else:
            raise SchemaExtractError(f"unsupported predicate operator: {operator!r}")
        expression = current if expression is None else expression & current
    return expression


def scan_arrow_batches(
    dataset_path: Path,
    *,
    columns: list[str] | None = None,
    predicates: list[tuple[str, str, Any]] | None = None,
    batch_size: int = 65_536,
):
    """Lazily scan Parquet with projection, predicate pushdown, and batches."""
    path = Path(dataset_path)
    if path.suffix.casefold() != ".parquet":
        raise SchemaExtractError("lazy Arrow scanning currently requires Parquet")
    if not isinstance(batch_size, int) or not 1 <= batch_size <= 1_000_000:
        raise SchemaExtractError("batch_size must be between 1 and 1,000,000")
    _guard_columnar_expansion(path, full_load_max_bytes(), columns=columns)
    import pyarrow.dataset as ds

    dataset = ds.dataset(path, format="parquet")
    scanner = dataset.scanner(
        columns=columns,
        filter=_arrow_filter_expression(predicates),
        batch_size=batch_size,
        use_threads=True,
    )
    return scanner.to_batches()


def _record_looks_like_header(record: list[str]) -> bool:
    """Heuristic: does a parsed CSV/TSV record look like a header row
    (column names) or a data row?

    Takes an already-parsed list of cells from ``csv.reader`` (which
    correctly handles quoted, multi-line fields) rather than raw
    bytes. True (header) when at least one cell can't be parsed as a
    number. Same edge-case posture as the previous bytes-based
    version:

    - Empty or single-cell record: treated as a header (a single-
      column file with a name like ``"id"`` is the common case).
    - Every cell is numeric (raw sensor dump etc.): treated as data,
      no header offset applied.
    """
    if not record:
        return True
    saw_any = False
    for cell in record:
        s = (cell or "").strip()
        if not s:
            continue
        saw_any = True
        try:
            float(s)
        except ValueError:
            return True
    if not saw_any:
        return True
    return False


def _gzip_has_header(path: Path, delimiter: str) -> bool:
    import csv
    import gzip

    with gzip.open(path, "rt", encoding="utf-8", errors="replace", newline="") as handle:
        first = next(csv.reader(handle, delimiter=delimiter), [])
    return _record_looks_like_header(first)


# ---------------------------------------------------------------------------
# Text-table dialect detection (encoding, delimiter, decimal mark)
# ---------------------------------------------------------------------------
#
# Real research CSVs are not the UTF-8 comma-separated ideal. European
# Excel exports ``;``-separated files with decimal commas; legacy SPSS
# and Access exports arrive in Latin-1/Windows-1252; Windows tools
# prepend BOMs or write UTF-16. Before this chokepoint, a Latin-1 file
# died with a UnicodeDecodeError and — far worse — a semicolon CSV
# parsed *silently* into one mashed column, producing a garbage schema
# with no error anywhere.
#
# One function decides (encoding, delimiter, decimal) from a bounded
# head sample, and every reader of .csv/.tsv goes through it, so the
# schema, the row-count audit, the full load and the profile cannot
# disagree about what the file says. All heuristics are deterministic
# (no chardet dependency) and total (Latin-1 decodes any byte string,
# so the fallback cannot itself fail).

_SNIFF_BYTES = 64 * 1024
_CSV_CANDIDATE_DELIMITERS = (",", ";", "\t", "|")


def text_table_params(path: Path, suffix: str) -> tuple[str, str, str]:
    """Return ``(encoding, delimiter, decimal)`` for a text table.

    ``suffix`` decides the delimiter search space: ``.tsv`` is tab by
    definition; ``.csv`` is sniffed among comma / semicolon / tab /
    pipe. Decimal commas are inferred only for semicolon files, where
    the pairing is the near-universal European convention — guessing
    decimal commas in a comma-separated file would be unfalsifiable
    from a sample and is not attempted.
    """
    try:
        with path.open("rb") as handle:
            raw = handle.read(_SNIFF_BYTES)
    except OSError:
        return "utf-8", ("\t" if suffix == ".tsv" else ","), "."

    # Encoding: BOMs first (they are unambiguous), then strict UTF-8,
    # then Latin-1 — which maps every byte, so this chain is total.
    if raw.startswith(b"\xef\xbb\xbf"):
        encoding = "utf-8-sig"
    elif raw.startswith(b"\xff\xfe") or raw.startswith(b"\xfe\xff"):
        encoding = "utf-16"
    else:
        try:
            raw.decode("utf-8")
            encoding = "utf-8"
        except UnicodeDecodeError:
            encoding = "latin-1"

    try:
        sample = raw.decode(encoding, errors="replace")
    except (UnicodeDecodeError, LookupError):
        sample = raw.decode("latin-1", errors="replace")

    if suffix == ".tsv":
        return encoding, "\t", "."

    lines = [ln for ln in sample.splitlines() if ln.strip()][:20]
    delimiter = ","
    if lines:
        best_count = 0
        for cand in _CSV_CANDIDATE_DELIMITERS:
            counts = [ln.count(cand) for ln in lines]
            # A real delimiter appears the same number of times on
            # every sampled line (quoted fields can break this for
            # pathological files; those fall back to comma, which is
            # pandas' behaviour today — no regression).
            if counts[0] > 0 and all(c == counts[0] for c in counts):
                if counts[0] > best_count:
                    best_count = counts[0]
                    delimiter = cand
    decimal = "."
    if delimiter == ";" and lines:
        body = "\n".join(lines[1:] or lines)
        fields = [f.strip() for ln in body.splitlines()
                  for f in ln.split(";")]
        has_comma_dec = any(_looks_decimal(f, ",") for f in fields)
        has_dot_dec = any(_looks_decimal(f, ".") for f in fields)
        if has_comma_dec and not has_dot_dec:
            decimal = ","
    return encoding, delimiter, decimal


def _looks_decimal(field: str, mark: str) -> bool:
    if mark not in field:
        return False
    left, _, right = field.partition(mark)
    return (left.lstrip("-").isdigit() and right.isdigit()
            and bool(left.lstrip("-")) and bool(right))


def _csv_has_header(path: Path, delimiter: str) -> bool:
    """Single source of truth for "does this CSV/TSV have a header?"
    so :func:`row_count` and the names_only fast paths agree.

    Without this shared peek, the two surfaces could disagree on a
    headerless numeric CSV: ``row_count`` (which runs its own header
    heuristic) reports N rows of data, while the fast-path extract
    (which used pandas' default ``header='infer'``, i.e. always treat
    row 1 as header) silently labels the data row as column names AND
    reports a row count that doesn't match what a full extract would
    return. The schema response then looks internally inconsistent
    (variable names look like data, observation count is one too many).

    Reads only the first parsed record — constant-time on any file
    size. Returns ``True`` (header present) on read errors so callers
    fall back to pandas' usual behaviour rather than guessing.
    """
    import csv

    try:
        suffix = path.suffix.lower()
        encoding, sniffed_delim, _dec = text_table_params(path, suffix)
        # Caller-supplied delimiter kept for backward compatibility;
        # the sniffed one wins for .csv so a semicolon file's header
        # peek parses the same record the readers will.
        if suffix == ".csv":
            delimiter = sniffed_delim
        with open(
            path, "r", encoding=encoding, errors="replace", newline="",
        ) as f:
            reader = csv.reader(f, delimiter=delimiter)
            for record in reader:
                return _record_looks_like_header(record)
    except OSError:
        return True
    # Empty file — no record at all. Default to "has header" so the
    # downstream pandas read uses its default and yields a zero-row
    # frame rather than failing on an unexpected option.
    return True


def row_count(dataset_path: Path) -> int | None:
    """Return the row count for the dataset at ``dataset_path`` without
    materialising the values where the format allows it.

    Used by ``submit_script``'s row-count audit (one comparison per
    multi-result call against the source dataset's N). The audit doesn't
    need any column data — just the row count — so we deliberately
    avoid the full ``load_data`` path which can take 60+ seconds on a
    multi-GB .dta.

    Per format:
    - ``.dta``: ``pyreadstat.read_dta(metadataonly=True).meta.number_rows``.
      0.5s on a 3 GB file vs ~60s for a full load.
    - ``.parquet``: ``pyarrow.parquet.ParquetFile(...).metadata.num_rows``.
      Reads only the footer.
    - ``.csv`` / ``.tsv``: byte-streamed line count, minus 1 if the
      first line looks like a header (any non-numeric token in the
      first row). The previous unconditional ``-1`` was wrong for
      headerless dumps (raw instrument files, anonymous panel data,
      log files renamed to .csv) — it produced an audit count off
      by one and the row-count check then false-flagged scripts
      that correctly counted the headerless row. Heuristic-only;
      a CSV whose header happens to be numeric (rare but possible
      — column names like ``"123"``) still gets misclassified, but
      the audit treats ``None`` and "off by 1" the same way (best-
      effort flag, not a gate).
    - ``.jsonl`` / ``.ndjson``: byte-streamed line count.
    - ``.rds``: no light path available without spinning up R; falls
      back to ``load_data`` and counts.

    Returns ``None`` on any failure — the audit is a best-effort signal,
    not a gate, so callers should treat ``None`` as "skip the check"
    rather than raise.
    """
    suffix = _logical_suffix(dataset_path)
    try:
        if suffix in _EXPLICIT_SELECTION_EXTENSIONS:
            return None
        if suffix == ".dta":
            import pyreadstat
            _df, meta = pyreadstat.read_dta(
                str(dataset_path), metadataonly=True,
            )
            n = meta.number_rows
            return int(n) if n is not None else None
        if suffix in (".sav", ".zsav", ".por", ".sas7bdat", ".xpt"):
            # Same metadata-only path as .dta. SPSS and sas7bdat carry
            # the row count in the header; XPORT does not
            # (``meta.number_rows`` is None there) — returning None is
            # the documented "skip the audit" signal, and a full read
            # just to audit a count would defeat the fast path.
            import pyreadstat
            reader: Any = {
                ".sav": pyreadstat.read_sav,
                ".zsav": pyreadstat.read_sav,
                ".por": pyreadstat.read_por,
                ".sas7bdat": pyreadstat.read_sas7bdat,
                ".xpt": pyreadstat.read_xport,
            }[suffix]
            _df, meta = reader(str(dataset_path), metadataonly=True)
            n = meta.number_rows
            return int(n) if n is not None else None
        if suffix == ".xlsx":
            # Streamed count via openpyxl read-only mode; minus one
            # for the assumed header row (extractor and load_data make
            # the same assumption, so the three views agree).
            from openpyxl import load_workbook
            wb = load_workbook(
                str(dataset_path), read_only=True, data_only=True
            )
            try:
                ws = wb.worksheets[0]
                n_rows = sum(1 for _ in ws.iter_rows(values_only=True))
            finally:
                wb.close()
            return max(0, n_rows - 1)
        if suffix == ".xls":
            import xlrd

            book = xlrd.open_workbook(str(dataset_path), on_demand=True)
            try:
                return max(0, int(book.sheet_by_index(0).nrows) - 1)
            finally:
                book.release_resources()
        if suffix == ".ods":
            # ODF has no stable cheap row-count metadata path in pandas.
            # This is guarded by the same full-load/zip expansion ceiling.
            return int(len(load_data(dataset_path)))
        if suffix == ".parquet":
            import pyarrow.parquet as pq
            return int(pq.ParquetFile(str(dataset_path)).metadata.num_rows)
        if suffix == ".feather":
            import pyarrow.feather as feather

            return int(feather.read_table(str(dataset_path), columns=[]).num_rows)
        if suffix in (".arrow", ".ipc"):
            with _open_arrow_ipc(dataset_path) as (reader, streaming):
                return sum(
                    batch.num_rows for batch in _arrow_batches(reader, streaming)
                )
        if suffix == ".orc":
            import pyarrow.orc as orc

            return int(orc.ORCFile(str(dataset_path)).nrows)
        if suffix == ".csv" or suffix == ".tsv":
            # Streamed parse via ``csv.reader`` so quoted fields with
            # embedded newlines count as ONE record, not multiple.
            # The previous byte-streamed line counter treated every
            # physical ``\n`` as a row boundary, so a valid CSV like
            #   id,note\n
            #   1,"hello\n
            #   world"\n
            # was counted as 3 lines (- 1 for header) = 2 rows even
            # though the actual analysis sees 1 row, false-flagging
            # the row-count audit. ``csv.reader`` honours RFC 4180
            # quoting and is memory-efficient (it streams the file
            # without materialising the whole thing). The header
            # heuristic operates on the parsed first record so it
            # works through quotes correctly.
            import csv
            delimiter = "," if suffix == ".csv" else "\t"
            n_records = 0
            first_record: list[str] | None = None
            with open(
                dataset_path, "r", encoding="utf-8",
                errors="replace", newline="",
            ) as f:
                csv_rows = csv.reader(f, delimiter=delimiter)
                for record in csv_rows:
                    if first_record is None:
                        first_record = record
                    n_records += 1
            if n_records == 0:
                return 0
            has_header = _record_looks_like_header(first_record or [])
            return max(0, n_records - 1 if has_header else n_records)
        if suffix in {".csv.gz", ".tsv.gz"}:
            import csv
            import gzip

            delimiter = "\t" if suffix == ".tsv.gz" else ","
            n_records = 0
            gzip_first_record: list[str] | None = None
            with gzip.open(
                dataset_path, "rt", encoding="utf-8", errors="replace", newline="",
            ) as handle:
                for record in csv.reader(handle, delimiter=delimiter):
                    if gzip_first_record is None:
                        gzip_first_record = record
                    n_records += 1
            return max(
                0, n_records - 1
                if _record_looks_like_header(gzip_first_record or []) else n_records,
            )
        if suffix in (".jsonl", ".ndjson"):
            n_lines = 0
            with open(dataset_path, "rb") as f:
                for line in f:
                    if line.strip():
                        n_lines += 1
            return n_lines
        if suffix in {".jsonl.gz", ".ndjson.gz"}:
            import gzip

            with gzip.open(dataset_path, "rb") as handle:
                return sum(1 for line in handle if line.strip())
        if suffix == ".json":
            return len(load_data(dataset_path))
        if suffix in {".rds", ".rda", ".rdata"}:
            # No metadata-only path; fall back to full load.
            return int(len(load_data(dataset_path)))
    except Exception:  # noqa: BLE001 — metadata estimation is best-effort
        return None
    return None


# Formats and depths that require a whole-file load. ``extract()`` must apply
# the memory-safety guard before dispatching any of these paths.
#   - .rds: pyreadr has no partial-read API. ALWAYS a full load,
#     regardless of depth.
#   - .dta/.sav/.zsav/.sas7bdat/.xpt: pyreadstat's ``metadataonly=True``
#     is cheap and used for every depth except the deepest
#     (names_types_labels_summary), which loads real values.
#   - .xlsx: the deepest depth reads the whole sheet; shallower depths
#     pass ``nrows=200`` to openpyxl's streaming (``read_only=True``)
#     reader, which is bounded to (approximately) the first 200 rows'
#     worth of decompression regardless of total sheet size.
#   - .csv/.tsv/.parquet/.jsonl/.ndjson: ``names_only`` is a cheap
#     header/footer/streaming-key peek for all four; every other depth
#     does a full ``pd.read_*``.
def _extract_will_full_load(suffix: str, depth: str) -> bool:
    if suffix in {".rds", ".rda", ".rdata"}:
        return True
    if suffix in (
        ".dta", ".sav", ".zsav", ".por", ".sas7bdat", ".xpt",
        ".xlsx", ".xls", ".ods",
    ):
        return depth == "names_types_labels_summary"
    if suffix in (
        ".csv", ".tsv", ".parquet", ".feather", ".arrow", ".ipc", ".orc",
        ".jsonl", ".ndjson",
    ):
        return depth != "names_only"
    if suffix in {".csv.gz", ".tsv.gz", ".jsonl.gz", ".ndjson.gz"}:
        return True
    if suffix == ".json":
        return True
    return False


def extract(
    dataset_path: Path, depth: str, *, sheet: str | int | None = None,
) -> dict[str, Any]:
    """Return a structured schema summary for the file at `dataset_path`.

    Dispatches by suffix. Raises ValueError for unsupported formats or
    invalid depth; lets underlying library errors propagate (tool layer
    catches them and returns a policy-shaped error payload).

    ``sheet`` selects a worksheet for ``.xlsx`` (see ``load_data``);
    ignored for every other format.
    """
    if depth not in _VALID_DEPTHS:
        raise SchemaExtractError(
            f"invalid depth: {depth!r}; valid: {sorted(_VALID_DEPTHS)}"
        )
    suffix = _logical_suffix(dataset_path)
    if suffix in _EXPLICIT_SELECTION_EXTENSIONS:
        raise SchemaExtractError(
            f"{suffix} requires an explicit object/variable/layer selection; "
            "materialize it through Sift's isolated format-selection workflow first"
        )
    if _extract_will_full_load(suffix, depth):
        _guard_full_load(dataset_path)
    if suffix == ".dta":
        return _extract_stata(dataset_path, depth)
    if suffix in (".sav", ".zsav", ".por"):
        return _extract_spss(dataset_path, depth)
    if suffix == ".sas7bdat":
        return _extract_sas(dataset_path, depth)
    if suffix == ".xpt":
        return _extract_xport(dataset_path, depth)
    if suffix in (".xlsx", ".xls", ".ods"):
        return _extract_excel(dataset_path, depth, sheet=sheet)
    if suffix in {".rds", ".rda", ".rdata"}:
        return _extract_rds(dataset_path, depth)
    if suffix == ".csv":
        return _extract_csv(dataset_path, depth)
    if suffix == ".tsv":
        return _extract_tsv(dataset_path, depth)
    if suffix == ".parquet":
        return _extract_parquet(dataset_path, depth)
    if suffix in (".feather", ".arrow", ".ipc", ".orc"):
        return _extract_arrow(dataset_path, depth)
    if suffix in (".jsonl", ".ndjson"):
        return _extract_jsonl(dataset_path, depth)
    if suffix == ".json":
        df = load_data(dataset_path)
        return _extract_from_pandas(
            df,
            depth=depth,
            dataset_name=dataset_path.name,
            file_type="json",
        )
    if suffix in {".csv.gz", ".tsv.gz", ".jsonl.gz", ".ndjson.gz"}:
        df = load_data(dataset_path)
        return _extract_from_pandas(
            df, depth=depth, dataset_name=dataset_path.name,
            file_type=suffix.removeprefix(".").replace(".", "_"),
        )
    raise SchemaExtractError(
        f"unsupported format: {suffix!r}. Sift currently reads "
        ".csv, .tsv, .dta (Stata), .sav/.zsav/.por (SPSS), .sas7bdat/.xpt "
        "(SAS), .xlsx/.xls/.ods (spreadsheets), .rds/.rda/.RData (R), .parquet, "
        ".feather, .arrow/.ipc, .orc, .json, .jsonl, .ndjson."
    )


# ---------------------------------------------------------------------------
# Stata — .dta via pyreadstat
# ---------------------------------------------------------------------------

# Per-variable cap on emitted value-label entries. A codebook-heavy
# .dta (e.g. an industry classification with thousands of NAICS codes)
# would otherwise pour every label into the schema response, blowing
# context and creating a large data-origin text channel even after
# per-string sanitization. We surface the count and a hint so the
# model can request the full codebook through a different path if
# it actually needs it.
_MAX_VALUE_LABELS_PER_VAR = 50
# Total cap across all variables in one schema response, so a file
# with many medium-sized label sets can't blow the budget either.
_MAX_VALUE_LABELS_TOTAL = 500


class SchemaExtractError(ValueError):
    """Raised by ``schema.extract`` for parser-OWNED validation errors.

    These messages are crafted by this module (unsupported format,
    invalid depth, .rds-without-dataframe, etc.) and are safe to
    forward verbatim — they do not embed row content or library
    diagnostics. The tool layer relies on the class to distinguish
    them from data-leak-prone pandas / pyreadstat / pyreadr
    exceptions (some of which are also ``ValueError`` subclasses,
    notably ``pandas.errors.ParserError`` whose message quotes the
    offending CSV row).
    """


class DatasetTooLargeError(SchemaExtractError):
    """Raised when a full in-memory load would exceed the size ceiling.

    Subclasses ``SchemaExtractError`` so existing ``except`` sites keep
    working; callers that want the specific "too large" case (to render
    a better hint) can catch this type.
    """


def _extract_stata(path: Path, depth: str) -> dict[str, Any]:
    import pyreadstat

    return _extract_pyreadstat(
        path, depth, pyreadstat.read_dta, file_type="stata",
    )


def _extract_spss(path: Path, depth: str) -> dict[str, Any]:
    """SPSS ``.sav`` / ``.zsav`` / portable ``.por``.

    Same libreadstat backend and identical metadata container as the
    Stata path: ``read_sav`` fills ``column_labels``,
    ``variable_to_label`` / ``value_labels`` and
    ``readstat_variable_types`` exactly as ``read_dta`` does (verified
    empirically against written files), so the whole extraction —
    including the value-label budget and injection sanitisation —
    is shared rather than re-implemented.
    """
    import pyreadstat

    reader = pyreadstat.read_por if path.suffix.lower() == ".por" else pyreadstat.read_sav
    file_type = "spss_por" if path.suffix.lower() == ".por" else "spss"
    return _extract_pyreadstat(path, depth, reader, file_type=file_type)


def _extract_sas(path: Path, depth: str) -> dict[str, Any]:
    """SAS ``.sas7bdat`` — native SAS datasets."""
    import pyreadstat

    return _extract_pyreadstat(
        path, depth, pyreadstat.read_sas7bdat, file_type="sas",
    )


def _extract_xport(path: Path, depth: str) -> dict[str, Any]:
    """SAS Transport ``.xpt`` — the FDA submission format."""
    import pyreadstat

    return _extract_pyreadstat(
        path, depth, pyreadstat.read_xport, file_type="sas_xport",
    )


def _extract_pyreadstat(
    path: Path,
    depth: str,
    reader: Any,
    *,
    file_type: str = "stata",
) -> dict[str, Any]:
    wants_summary = depth == "names_types_labels_summary"
    # metadataonly is fast and avoids loading any values. We only load the
    # DataFrame when we need summary stats.
    if wants_summary:
        df, meta = reader(str(path))
    else:
        df, meta = reader(str(path), metadataonly=True)

    variables: list[dict[str, Any]] = []
    labels_emitted_total = 0
    # Variable names originate in the data file and are forwarded to
    # Claude — pass through the injection defense. Computed as one
    # batch (not per-iteration) so two distinct column names that
    # happen to share a >40-char prefix get disambiguated instead of
    # silently rendering as two identical-looking "name" entries a
    # later request_data/submit_script call could resolve ambiguously.
    safe_names = safe_keys_sequence(meta.column_names)
    for idx, name in enumerate(meta.column_names):
        safe_name = safe_names[idx]
        var: dict[str, Any] = {"name": safe_name}

        if depth != "names_only":
            var["type"] = _stata_type(meta, name)

        if depth in ("names_types_labels", "names_types_labels_summary"):
            col_labels = meta.column_labels or []
            if idx < len(col_labels):
                label = col_labels[idx]
                if label:
                    # Variable labels are free-text — the longest injection
                    # surface in a typical .dta. Sanitize aggressively.
                    var["label"] = safe_text(str(label))
            # Value labels, if this column is tied to a label set. Both
            # the codes (keys) and labels (values) originate in the data,
            # so we sanitize each entry AND cap the count: a codebook-
            # heavy file (industry classifications, geographic codes)
            # could otherwise emit thousands of labels per variable
            # and tens of thousands across the file, spending the
            # context window and providing a wide data-origin text
            # channel.
            label_set = meta.variable_to_label.get(name)
            if label_set and label_set in meta.value_labels:
                raw = meta.value_labels[label_set]
                total_in_set = len(raw)
                budget_remaining = max(
                    0, _MAX_VALUE_LABELS_TOTAL - labels_emitted_total
                )
                effective_cap = min(
                    _MAX_VALUE_LABELS_PER_VAR, budget_remaining
                )
                # Stable insertion order from pyreadstat; take the
                # first ``effective_cap`` entries so repeated calls
                # against the same file return the same view.
                items = list(raw.items())[:effective_cap]
                # Two distinct raw codes can truncate to the SAME
                # safe_key when both are longer than the 40-char cap
                # (or share a long common prefix) — a naive
                # dict-comprehension here would let the second
                # silently overwrite the first, dropping a value
                # label with no error and no count mismatch a caller
                # could notice. safe_keys_sequence detects and
                # disambiguates that case instead of a plain
                # {safe_key(k): ... for k, v in items} collapsing it.
                item_keys = safe_keys_sequence([k for k, _ in items])
                var["value_labels"] = {
                    safe_k: safe_text(str(v))
                    for safe_k, (_, v) in zip(item_keys, items)
                }
                labels_emitted_total += len(items)
                if total_in_set > len(items):
                    var["value_labels_total"] = total_in_set
                    var["value_labels_truncated"] = True

        # Use the ORIGINAL name as the key when reading df (to match pandas),
        # but report the SANITIZED name to Claude.
        if wants_summary and df is not None and name in df.columns:
            series = df[name]
            n_obs = int(len(series))
            raw_na = int(series.isna().sum())
            var["na_count"] = _suppress_rare_count(
                raw_na, n_obs, _SCHEMA_SUMMARY_THRESHOLD,
            )
            if var.get("type") == _TYPE_CATEGORICAL:
                var["distinct_count"] = int(series.nunique(dropna=True))

        variables.append(var)

    return {
        "status": "ok",
        # Filename crosses to Claude as text, so it's a prompt-injection
        # surface: a file named with embedded newlines or fake system
        # markers would land in the model's context verbatim. safe_text
        # strips control chars, flattens whitespace, and caps length —
        # same chokepoint we apply to variable labels above. "filename
        # only — no path injection surface" was the old comment; it
        # covered path-traversal but NOT prompt injection.
        "dataset": safe_text(path.name),
        "file_type": file_type,
        "depth": depth,
        # Some metadata-only readers (notably SAS XPORT) do not expose a
        # row count. ``null`` means "unknown" throughout the schema contract;
        # it must never be misreported as zero or crash via ``int(None)``.
        "observation_count": (
            int(meta.number_rows) if meta.number_rows is not None else None
        ),
        "variables": variables,
    }


def _stata_type(meta: Any, name: str) -> str:
    """Map pyreadstat's Stata metadata to our coarse type taxonomy."""
    # Value labels → treat as categorical (regardless of underlying numeric
    # encoding — that's how Stata users think about them).
    if meta.variable_to_label.get(name):
        return _TYPE_CATEGORICAL
    typ = (meta.readstat_variable_types or {}).get(name, "").lower()
    if typ == "string":
        return _TYPE_STRING
    if typ in ("int8", "int16", "int32"):
        return _TYPE_INTEGER
    if typ in ("float", "double"):
        return _TYPE_NUMERIC
    return _TYPE_UNKNOWN


# ---------------------------------------------------------------------------
# Excel — .xlsx via pandas/openpyxl
# ---------------------------------------------------------------------------
#
# Scope decisions, stated rather than implied:
# - First worksheet only. Multi-sheet workbooks are workbooks, not
#   datasets; picking sheet 0 is predictable, and the researcher can
#   save a specific sheet as its own file when it matters. The
#   response names the sheet that was read so nothing is silent.
# - Header row assumed. That is the overwhelming Excel convention;
#   a headerless numeric sheet reads with its first row as names,
#   which is visible immediately in the schema rather than corrupting
#   quietly downstream.
# - No value labels — the format has none.

def list_excel_sheets(path: Path) -> list[str]:
    """Worksheet names in an ``.xlsx`` workbook, in file order.

    Metadata-only: ``pandas.ExcelFile`` (openpyxl-backed) parses the
    workbook's sheet directory without materializing any cell data,
    so this is cheap even on a large workbook. Used by the Data
    panel to offer a sheet picker, and — indirectly, via
    ``_extract_xlsx`` below — surfaced to the model in every
    ``get_schema`` response so it knows what else is available
    without a separate round trip.
    """
    import pandas as pd

    with pd.ExcelFile(path) as xl:
        return list(xl.sheet_names)


def _extract_excel(
    path: Path, depth: str, *, sheet: str | int | None = None,
) -> dict[str, Any]:
    import pandas as pd

    sheet_to_read = 0 if sheet is None else sheet
    wants_summary = depth == "names_types_labels_summary"
    if wants_summary:
        df = pd.read_excel(path, sheet_name=sheet_to_read)
    else:
        # Bounded head read: enough rows to type columns without a
        # full load at shallow depths.
        df = pd.read_excel(path, sheet_name=sheet_to_read, nrows=200)
    # Shared pandas extraction path — same typing, same sanitisation,
    # same summary semantics as CSV/TSV, so Excel cannot drift.
    result = _extract_from_pandas(
        df, depth=depth, dataset_name=path.name,
        file_type=path.suffix.lower().removeprefix("."),
    )
    # Worksheet scope is a real decision; say it, don't imply it. The
    # model reads this to know which sheet its own ``pd.read_excel`` /
    # ``readxl::read_excel`` / ``import excel`` call in a submitted
    # script needs to reference to match what get_schema described.
    result["sheet_read"] = sheet_to_read
    try:
        result["available_sheets"] = list_excel_sheets(path)
    except Exception:  # noqa: BLE001 — advisory only, never block schema
        pass
    if not wants_summary:
        # The head read can't know the true row count; never report
        # the sample size as if it were one.
        result.pop("observation_count", None)
    return result


# ---------------------------------------------------------------------------
# R — .rds / .rda / .RData via pyreadr
# ---------------------------------------------------------------------------

def _read_single_r_dataframe(path: Path, object_name: str | None = None) -> Any:
    """Load one unambiguous R data frame.

    An R workspace may contain several objects. Choosing the first would make
    schema, row counts, and analysis depend on serialization order, which is
    unacceptable for reproducible research. Require exactly one data-frame
    object and ask the researcher to split ambiguous workspaces in R.
    """
    import pyreadr

    result = pyreadr.read_r(str(path))
    frames = {
        str(key): value for key, value in result.items()
        if hasattr(value, "columns")
    }
    suffix = path.suffix
    if not result:
        raise SchemaExtractError(f"{suffix} file contains no objects: {path}")
    if object_name is not None:
        selected = frames.get(str(object_name))
        if selected is None:
            raise SchemaExtractError(
                f"selected R data-frame object {object_name!r} was not found"
            )
        return selected
    if len(result) != 1 or len(frames) != 1:
        raise SchemaExtractError(
            f"{suffix} file must contain exactly one data frame for an "
            "unambiguous analysis; found {len(frames)} data frames across "
            f"{len(result)} readable objects. Save the intended data frame "
            "as a separate .rds file or .RData workspace."
        )
    return next(iter(frames.values()))


def _extract_rds(path: Path, depth: str) -> dict[str, Any]:
    df = _read_single_r_dataframe(path)

    return _extract_from_pandas(
        df,
        depth=depth,
        dataset_name=path.name,
        file_type="r_rds" if path.suffix.lower() == ".rds" else "r_workspace",
    )


# ---------------------------------------------------------------------------
# names_only fast path — column names only, no full data load
# ---------------------------------------------------------------------------

def _names_only_payload(
    column_names: list[str], path: Path, file_type: str,
) -> dict[str, Any]:
    """Build a names_only response from a list of column names and the
    path. Used by the CSV/TSV/Parquet/JSONL fast paths to avoid
    loading the whole dataset when the model only asked for the
    variable list.

    Observation count comes from ``row_count(path)`` (metadata- or
    streaming-only for these formats). A ``None`` return — meaning
    "the light path couldn't compute the count" — surfaces to the
    model as ``observation_count: null`` so it can tell "I don't
    know" apart from "the dataset has zero rows". Previously we
    coerced the None to 0, which made a parquet footer read failure
    or a mid-stream CSV encoding error look like a factual empty
    dataset.
    """
    obs = row_count(path)
    # Batched (not per-column safe_key) so two column names sharing a
    # >40-char prefix get disambiguated rather than rendering as two
    # identical-looking entries — see safe_keys_sequence's docstring.
    variables = [{"name": n} for n in safe_keys_sequence(column_names)]
    return {
        "status": "ok",
        # See _extract_stata for the prompt-injection rationale on
        # filenames passing through safe_text.
        "dataset": safe_text(path.name),
        "file_type": file_type,
        "depth": "names_only",
        "observation_count": int(obs) if obs is not None else None,
        "variables": variables,
    }


# ---------------------------------------------------------------------------
# CSV — pandas
# ---------------------------------------------------------------------------

def _extract_csv(path: Path, depth: str) -> dict[str, Any]:
    import pandas as pd

    if depth == "names_only":
        # Fast path: ``nrows=0`` makes pandas read only the header
        # row and return an empty DataFrame with the correct columns.
        # For a multi-GB CSV this is constant-time + a single read of
        # the first line, vs a full pass that materialises every
        # column in memory for type inference. A bare "what columns
        # does this dataset have" call should not OOM the app.
        #
        # Honour the SAME header heuristic ``row_count`` uses so the
        # two outputs stay consistent: a headerless numeric CSV
        # (``1,2,3\n4,5,6``) was reporting variable names ``"1","2","3"``
        # alongside observation_count=2, but a full extract would
        # have consumed row 1 as the header and reported only 1
        # data row. With the shared peek, both surfaces agree:
        # headerless files come back with ``0,1,2…`` placeholder
        # names and the full row count.
        enc, sep, _dec = text_table_params(path, ".csv")
        has_header = _csv_has_header(path, sep)
        header_df = pd.read_csv(
            path, sep=sep, encoding=enc,
            header=0 if has_header else None, nrows=0,
        )
        return _names_only_payload(
            list(header_df.columns), path, "csv",
        )
    # low_memory=False gives a single-pass type inference — more accurate
    # for columns where the type isn't obvious from the first chunk. For
    # genuinely huge CSVs this will be slow; step 7 can add streaming.
    enc, sep, dec = text_table_params(path, ".csv")
    has_header = _csv_has_header(path, sep)
    df = pd.read_csv(
        path, sep=sep, encoding=enc, decimal=dec,
        header=0 if has_header else None, low_memory=False,
    )
    return _extract_from_pandas(
        df, depth=depth, dataset_name=path.name, file_type="csv"
    )


# ---------------------------------------------------------------------------
# TSV — same as CSV, tab-separated
# ---------------------------------------------------------------------------

def _extract_tsv(path: Path, depth: str) -> dict[str, Any]:
    import pandas as pd

    if depth == "names_only":
        # See _extract_csv for the fast-path rationale and the
        # header-heuristic alignment with ``row_count``.
        enc, _sep, _dec = text_table_params(path, ".tsv")
        has_header = _csv_has_header(path, "\t")
        header_df = pd.read_csv(
            path, sep="\t", encoding=enc,
            header=0 if has_header else None, nrows=0,
        )
        return _names_only_payload(
            list(header_df.columns), path, "tsv",
        )
    enc, _sep, _dec = text_table_params(path, ".tsv")
    has_header = _csv_has_header(path, "\t")
    df = pd.read_csv(
        path, sep="\t", encoding=enc,
        header=0 if has_header else None, low_memory=False,
    )
    return _extract_from_pandas(
        df, depth=depth, dataset_name=path.name, file_type="tsv"
    )


# ---------------------------------------------------------------------------
# Parquet — pyarrow-backed via pandas
# ---------------------------------------------------------------------------

def _extract_parquet(path: Path, depth: str) -> dict[str, Any]:
    """Parquet preserves column dtypes natively (unlike CSV, which the
    extractor has to infer). Schema extraction is therefore a thin
    wrapper around ``read_parquet``: pandas hands back a DataFrame
    whose dtypes already match what the writer set.

    ``names_only`` reads the Parquet footer without loading row data. Deeper
    requests require a full read and are protected by the dispatcher-level
    size guard before this function is called."""
    import pandas as pd

    if depth == "names_only":
        # Fast path: pyarrow exposes the Parquet schema in the file
        # footer (constant-time regardless of file size). The full-
        # load fallback below covers any depth that actually needs
        # column data.
        try:
            import pyarrow.parquet as pq
            pf = pq.ParquetFile(str(path))
            names = list(pf.schema_arrow.names)
            return _names_only_payload(names, path, "parquet")
        except Exception:  # noqa: BLE001 — fall through to full load
            pass
    df = pd.read_parquet(path)
    return _extract_from_pandas(
        df, depth=depth, dataset_name=path.name, file_type="parquet"
    )


def _extract_arrow(path: Path, depth: str) -> dict[str, Any]:
    """Arrow IPC/Feather/ORC with metadata-only names when possible."""
    suffix = path.suffix.lower()
    file_type = suffix.removeprefix(".")
    if depth == "names_only":
        try:
            if suffix == ".orc":
                import pyarrow.orc as orc

                names = list(orc.ORCFile(str(path)).schema.names)
            else:
                with _open_arrow_ipc(path) as (reader, _streaming):
                    names = list(reader.schema.names)
            return _names_only_payload(names, path, file_type)
        except Exception:  # noqa: BLE001 - guarded full-load fallback
            pass
    df = load_data(path)
    return _extract_from_pandas(
        df, depth=depth, dataset_name=path.name, file_type=file_type
    )


# ---------------------------------------------------------------------------
# Line-delimited JSON — .jsonl / .ndjson
# ---------------------------------------------------------------------------

# Hard cap on the number of distinct keys we'll collect from a JSONL
# names_only stream. A pathological file with millions of unique keys
# would otherwise turn the fast path into an unbounded-memory walk.
# 10_000 covers every legitimate tabular JSONL I've seen by orders of
# magnitude; beyond that the file isn't a "variable list" in any
# useful sense.
_JSONL_NAMES_KEY_CAP = 10_000


def _jsonl_column_union(path: Path) -> list[str]:
    """Stream a JSONL file and return the union of top-level keys,
    in first-seen order.

    Lines that aren't a JSON object (blank, malformed, top-level
    array / scalar) are skipped — the full-depth path via
    ``pd.read_json`` would coerce or raise on the same input, but we
    can't afford a raise here when the fast path is the only thing
    standing between the model and a missing-column lookup. Best
    effort: every well-formed object contributes its keys.
    """
    import json as _json

    seen: dict[str, None] = {}
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    record = _json.loads(line)
                except ValueError:
                    continue
                if not isinstance(record, dict):
                    continue
                for k in record.keys():
                    if k in seen:
                        continue
                    seen[k] = None
                    if len(seen) >= _JSONL_NAMES_KEY_CAP:
                        return list(seen.keys())
    except OSError:
        # Mirrors the full-depth path: a read failure surfaces through
        # the empty-names list, and ``row_count`` will return ``None``
        # so the payload says ``observation_count: null`` rather than
        # claiming a factual empty dataset.
        return []
    return list(seen.keys())


def _extract_jsonl(path: Path, depth: str) -> dict[str, Any]:
    """One JSON object per line. Each object is treated as a row;
    pandas infers column types from the union of keys.

    Top-level JSON arrays (a single ``[{...}, {...}]`` document) are
    NOT supported here. They're shape-arbitrary — a record could
    contain nested objects or arrays — and don't fit the tabular
    model the rest of the pipeline assumes. Researchers can convert
    with ``jq -c '.[]'`` if needed.
    """
    import pandas as pd

    if depth == "names_only":
        # Stream the file once and union the keys across every record.
        # JSONL is schema-less in the limit (each row may add new
        # keys), so a fast path that read only the first record could
        # silently miss columns — ``{"id":1}\n{"id":2,"treatment":1}``
        # would advertise only ``id`` at names_only while a deeper
        # depth (which loads everything via ``pd.read_json``) saw
        # both. That asymmetry broke the basic ``names_only`` contract
        # ("the list of variables in this dataset"); a model that
        # consulted names_only to discover columns then queried
        # ``treatment`` got "column not found".
        #
        # Streaming with ``json.loads`` per line and unioning keys
        # avoids materialising values or running pandas' type
        # inference — far cheaper than the full ``read_json`` we'd
        # otherwise use, while still seeing every column. Insertion
        # order is preserved so the names list reads first-seen-first.
        names = _jsonl_column_union(path)
        return _names_only_payload(names, path, "jsonl")
    df = pd.read_json(path, lines=True)
    return _extract_from_pandas(
        df, depth=depth, dataset_name=path.name, file_type="jsonl"
    )


# ---------------------------------------------------------------------------
# Generic pandas → schema
# ---------------------------------------------------------------------------

def _extract_from_pandas(
    df: Any,
    *,
    depth: str,
    dataset_name: str,
    file_type: str,
) -> dict[str, Any]:
    variables: list[dict[str, Any]] = []
    # Batched (not per-column safe_key) so two column names sharing a
    # >40-char prefix get disambiguated rather than rendering as two
    # identical-looking entries — see safe_keys_sequence's docstring.
    safe_names = safe_keys_sequence(df.columns)
    for idx, name in enumerate(df.columns):
        safe_name = safe_names[idx]
        var: dict[str, Any] = {"name": safe_name}
        if depth != "names_only":
            var["type"] = _pandas_type(df[name])
        # CSV / R data frames typically have no variable labels or value
        # labels metadata to expose at the names_types_labels depth. If
        # future formats (SPSS etc.) carry labels, we'll add them here.
        if depth == "names_types_labels_summary":
            series = df[name]
            n_obs = int(len(series))
            raw_na = int(series.isna().sum())
            var["na_count"] = _suppress_rare_count(
                raw_na, n_obs, _SCHEMA_SUMMARY_THRESHOLD,
            )
            if var.get("type") == _TYPE_CATEGORICAL:
                var["distinct_count"] = int(series.nunique(dropna=True))
        variables.append(var)
    return {
        "status": "ok",
        # See the identical note in _extract_stata_dta — filename is a
        # prompt-injection surface when echoed unsanitized.
        "dataset": safe_text(dataset_name),
        "file_type": file_type,
        "depth": depth,
        "observation_count": int(len(df)),
        "variables": variables,
    }


def _pandas_type(series: Any) -> str:
    """Map a pandas dtype to our coarse taxonomy."""
    import pandas as pd

    dtype = series.dtype
    if pd.api.types.is_bool_dtype(dtype):
        return _TYPE_BOOLEAN
    if pd.api.types.is_integer_dtype(dtype):
        return _TYPE_INTEGER
    if pd.api.types.is_float_dtype(dtype):
        return _TYPE_NUMERIC
    if pd.api.types.is_datetime64_any_dtype(dtype):
        return _TYPE_DATETIME
    if isinstance(dtype, pd.CategoricalDtype):
        return _TYPE_CATEGORICAL
    # Object columns in pandas are usually strings, but can also be
    # heterogeneous. We call them strings unless the unique count is small
    # enough to treat as categorical — matches how researchers actually
    # model them.
    if pd.api.types.is_object_dtype(dtype) or pd.api.types.is_string_dtype(dtype):
        try:
            n = len(series)
            if n == 0:
                return _TYPE_STRING
            nunique = series.nunique(dropna=True)
            # Heuristic: fewer than 20 distinct values and fewer than 5% of
            # total → categorical. Arbitrary but useful; researchers can
            # always override in their analysis.
            if nunique <= 20 and nunique <= max(1, n // 20):
                return _TYPE_CATEGORICAL
        except (TypeError, ValueError):
            pass
        return _TYPE_STRING
    return _TYPE_UNKNOWN
